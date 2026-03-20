# backend/app/api/routes/journal.py
"""
LuxQuant Terminal - Trade Journal Routes
=========================================
Full CRUD + Signal Prefill + Stats + AI Insights + Excel Export
Free for all authenticated users (no subscription check).
"""
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text, desc, asc, func
from typing import Optional, List
from datetime import datetime, timezone, timedelta
from collections import Counter
import io
import os
import httpx

from app.core.database import get_db
from app.core.redis import cache_get, cache_set
from app.api.deps import get_current_user
from app.models.user import User
from app.models.journal import TradeJournal
from app.schemas.journal import (
    JournalCreate, JournalUpdate, JournalResponse,
    JournalListResponse, JournalPrefillResponse,
    JournalStatsResponse, AIInsightResponse,
)

router = APIRouter(prefix="/journal", tags=["Trade Journal"])

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
GEMINI_MODEL = "gemini-2.5-flash"
GEMINI_URL = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"


# ════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════

def calc_pnl(entry: float, exit_price: float, size_usd: float,
             leverage: float, fees: float, direction: str) -> dict:
    """Calculate PnL from trade params."""
    if not exit_price or not entry or not size_usd:
        return {"pnl_usd": None, "pnl_pct": None}

    if direction == "short":
        raw_pnl = (entry - exit_price) / entry * size_usd * leverage
    else:
        raw_pnl = (exit_price - entry) / entry * size_usd * leverage

    pnl_usd = round(raw_pnl - fees, 2)
    pnl_pct = round((pnl_usd / size_usd) * 100, 2) if size_usd else 0
    return {"pnl_usd": pnl_usd, "pnl_pct": pnl_pct}


def calc_rr(entry: float, exit_price: float, sl: float, direction: str) -> Optional[float]:
    """Calculate risk:reward ratio."""
    if not exit_price or not entry or not sl:
        return None
    risk = abs(entry - sl)
    if risk == 0:
        return None
    reward = abs(exit_price - entry)
    return round(reward / risk, 2)


def derive_status(pnl_usd: Optional[float]) -> str:
    if pnl_usd is None:
        return "open"
    if pnl_usd > 0:
        return "closed_win"
    if pnl_usd < 0:
        return "closed_loss"
    return "breakeven"


# ════════════════════════════════════════════
# 1. PREFILL FROM SIGNAL
# ════════════════════════════════════════════

@router.get("/prefill/{signal_id}", response_model=JournalPrefillResponse)
async def prefill_from_signal(
    signal_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Auto-fill journal data from an existing LuxQuant signal."""
    row = db.execute(
        text("""
            SELECT s.pair, s.entry, s.target1, s.target2, s.target3, s.target4,
                   s.stop1, s.risk_level, s.status
            FROM signals s
            WHERE s.signal_id = :sid
        """),
        {"sid": signal_id},
    ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Signal not found")

    # Derive real status from signal_updates
    real_status_row = db.execute(
        text("""
            SELECT update_type FROM signal_updates
            WHERE signal_id = :sid
            ORDER BY CASE
                WHEN LOWER(update_type) LIKE '%tp4%' THEN 5
                WHEN LOWER(update_type) LIKE '%tp3%' THEN 4
                WHEN LOWER(update_type) LIKE '%tp2%' THEN 3
                WHEN LOWER(update_type) LIKE '%tp1%' THEN 2
                WHEN LOWER(update_type) LIKE '%stop%' OR LOWER(update_type) LIKE '%sl%' THEN 1
                ELSE 0
            END DESC
            LIMIT 1
        """),
        {"sid": signal_id},
    ).fetchone()

    signal_status = real_status_row[0] if real_status_row else (row[8] or "open")

    context = {
        "signal_status": signal_status,
        "risk_level": row[7],
    }

    return JournalPrefillResponse(
        pair=row[0],
        planned_entry=row[1],
        planned_tp1=row[2],
        planned_tp2=row[3],
        planned_tp3=row[4],
        planned_tp4=row[5],
        planned_sl=row[6],
        risk_level=row[7],
        signal_status=signal_status,
        context_snapshot=context,
    )


# ════════════════════════════════════════════
# 2. CRUD
# ════════════════════════════════════════════

@router.post("/", response_model=JournalResponse, status_code=201)
async def create_journal(
    data: JournalCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Create a new journal entry."""
    pnl = calc_pnl(
        data.actual_entry, data.actual_exit,
        data.position_size_usd or 0, data.leverage, data.fees_usd, data.direction,
    )
    rr = calc_rr(data.actual_entry, data.actual_exit, data.planned_sl, data.direction)
    st = derive_status(pnl["pnl_usd"]) if data.actual_exit else "open"

    # Build context snapshot if signal linked
    context = {}
    if data.signal_id:
        prefill = await prefill_from_signal(data.signal_id, current_user, db)
        context = prefill.context_snapshot

    journal = TradeJournal(
        user_id=current_user.id,
        signal_id=data.signal_id,
        pair=data.pair.upper(),
        direction=data.direction,
        status=st,
        planned_entry=data.planned_entry,
        planned_tp1=data.planned_tp1,
        planned_tp2=data.planned_tp2,
        planned_tp3=data.planned_tp3,
        planned_tp4=data.planned_tp4,
        planned_sl=data.planned_sl,
        actual_entry=data.actual_entry,
        actual_exit=data.actual_exit,
        leverage=data.leverage,
        position_size_usd=data.position_size_usd,
        fees_usd=data.fees_usd,
        pnl_usd=pnl["pnl_usd"],
        pnl_pct=pnl["pnl_pct"],
        rr_ratio=rr,
        emotions=data.emotions,
        strategy_tags=data.strategy_tags,
        confluence_tags=data.confluence_tags,
        mistakes=data.mistakes,
        notes=data.notes,
        chart_before_url=data.chart_before_url,
        chart_after_url=data.chart_after_url,
        tradingview_link=data.tradingview_link,
        context_snapshot=context,
        entry_at=data.entry_at or datetime.now(timezone.utc),
        exit_at=data.exit_at,
    )

    db.add(journal)
    db.commit()
    db.refresh(journal)
    return journal


@router.get("/", response_model=JournalListResponse)
async def list_journals(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    pair: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    strategy: Optional[str] = None,
    signal_linked: Optional[bool] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort_by: str = "entry_at",
    sort_order: str = "desc",
):
    """List all journal entries for current user with filters."""
    q = db.query(TradeJournal).filter(TradeJournal.user_id == current_user.id)

    if pair:
        q = q.filter(TradeJournal.pair == pair.upper())
    if status_filter:
        q = q.filter(TradeJournal.status == status_filter)
    if strategy:
        q = q.filter(TradeJournal.strategy_tags.any(strategy))
    if signal_linked is True:
        q = q.filter(TradeJournal.signal_id.isnot(None))
    elif signal_linked is False:
        q = q.filter(TradeJournal.signal_id.is_(None))
    if date_from:
        q = q.filter(TradeJournal.entry_at >= date_from)
    if date_to:
        q = q.filter(TradeJournal.entry_at <= date_to)

    sort_col = getattr(TradeJournal, sort_by, TradeJournal.entry_at)
    q = q.order_by(desc(sort_col) if sort_order == "desc" else asc(sort_col))

    items = q.all()
    return JournalListResponse(items=items, total=len(items))


@router.get("/{journal_id}", response_model=JournalResponse)
async def get_journal(
    journal_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get a single journal entry."""
    j = db.query(TradeJournal).filter(
        TradeJournal.id == journal_id,
        TradeJournal.user_id == current_user.id,
    ).first()
    if not j:
        raise HTTPException(status_code=404, detail="Journal entry not found")
    return j


@router.put("/{journal_id}", response_model=JournalResponse)
async def update_journal(
    journal_id: int,
    data: JournalUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Update/close a journal entry."""
    j = db.query(TradeJournal).filter(
        TradeJournal.id == journal_id,
        TradeJournal.user_id == current_user.id,
    ).first()
    if not j:
        raise HTTPException(status_code=404, detail="Journal entry not found")

    update_data = data.model_dump(exclude_unset=True)
    for key, val in update_data.items():
        setattr(j, key, val)

    # Recalculate PnL if exit changed
    exit_price = data.actual_exit if data.actual_exit is not None else j.actual_exit
    entry_price = data.actual_entry if data.actual_entry is not None else j.actual_entry
    size = data.position_size_usd if data.position_size_usd is not None else j.position_size_usd
    lev = data.leverage if data.leverage is not None else j.leverage
    fees = data.fees_usd if data.fees_usd is not None else j.fees_usd
    direction = data.direction if data.direction is not None else j.direction

    if exit_price:
        pnl = calc_pnl(entry_price, exit_price, size or 0, lev, fees, direction)
        j.pnl_usd = pnl["pnl_usd"]
        j.pnl_pct = pnl["pnl_pct"]
        j.rr_ratio = calc_rr(entry_price, exit_price, j.planned_sl, direction)
        if data.status is None:
            j.status = derive_status(pnl["pnl_usd"])

    db.commit()
    db.refresh(j)
    return j


@router.delete("/{journal_id}")
async def delete_journal(
    journal_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Delete a journal entry."""
    j = db.query(TradeJournal).filter(
        TradeJournal.id == journal_id,
        TradeJournal.user_id == current_user.id,
    ).first()
    if not j:
        raise HTTPException(status_code=404, detail="Journal entry not found")
    db.delete(j)
    db.commit()
    return {"detail": "Journal entry deleted"}


# ════════════════════════════════════════════
# 3. STATS / ANALYTICS
# ════════════════════════════════════════════

@router.get("/stats/overview", response_model=JournalStatsResponse)
async def get_journal_stats(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """Aggregated analytics from journal data."""
    q = db.query(TradeJournal).filter(TradeJournal.user_id == current_user.id)
    if date_from:
        q = q.filter(TradeJournal.entry_at >= date_from)
    if date_to:
        q = q.filter(TradeJournal.entry_at <= date_to)

    entries = q.order_by(TradeJournal.entry_at.asc()).all()

    if not entries:
        return JournalStatsResponse(
            total_trades=0, open_trades=0, closed_trades=0,
            wins=0, losses=0, breakeven=0, win_rate=0,
            total_pnl_usd=0, avg_pnl_usd=0, avg_rr=0,
            best_trade_pnl=0, worst_trade_pnl=0,
        )

    closed = [e for e in entries if e.status != "open"]
    wins = [e for e in closed if e.status == "closed_win"]
    losses = [e for e in closed if e.status == "closed_loss"]
    be = [e for e in closed if e.status == "breakeven"]

    total_pnl = sum(e.pnl_usd or 0 for e in closed)
    win_rate = (len(wins) / len(closed) * 100) if closed else 0
    avg_pnl = total_pnl / len(closed) if closed else 0
    rr_vals = [e.rr_ratio for e in closed if e.rr_ratio]
    avg_rr = sum(rr_vals) / len(rr_vals) if rr_vals else 0

    best = max(closed, key=lambda e: e.pnl_usd or 0) if closed else None
    worst = min(closed, key=lambda e: e.pnl_usd or 0) if closed else None

    # Most traded pair
    pair_counts = Counter(e.pair for e in entries)
    most_traded = pair_counts.most_common(1)[0][0] if pair_counts else None

    # Avg confidence/fomo for wins vs losses
    def avg_emotion(group, key):
        vals = [e.emotions.get(key) for e in group if e.emotions and e.emotions.get(key) is not None]
        return round(sum(vals) / len(vals), 1) if vals else None

    # Most common mistake
    all_mistakes = []
    for e in entries:
        if e.mistakes:
            all_mistakes.extend(e.mistakes)
    mistake_counts = Counter(all_mistakes)
    most_common_mistake = mistake_counts.most_common(1)[0][0] if mistake_counts else None

    # Win rate by strategy
    wr_by_strat = {}
    strat_counter = {}
    for e in closed:
        for tag in (e.strategy_tags or []):
            if tag not in strat_counter:
                strat_counter[tag] = {"wins": 0, "total": 0}
            strat_counter[tag]["total"] += 1
            if e.status == "closed_win":
                strat_counter[tag]["wins"] += 1
    for tag, c in strat_counter.items():
        wr_by_strat[tag] = {
            "win_rate": round(c["wins"] / c["total"] * 100, 1) if c["total"] else 0,
            "total": c["total"],
        }

    # Most profitable strategy
    best_strat = max(wr_by_strat, key=lambda k: wr_by_strat[k]["win_rate"]) if wr_by_strat else None

    # Win rate by emotion (mood)
    wr_by_emotion = {}
    for e in closed:
        mood = (e.emotions or {}).get("mood")
        if mood:
            if mood not in wr_by_emotion:
                wr_by_emotion[mood] = {"wins": 0, "total": 0}
            wr_by_emotion[mood]["total"] += 1
            if e.status == "closed_win":
                wr_by_emotion[mood]["wins"] += 1
    for mood, c in wr_by_emotion.items():
        wr_by_emotion[mood] = {
            "win_rate": round(c["wins"] / c["total"] * 100, 1) if c["total"] else 0,
            "total": c["total"],
        }

    # PnL by day of week
    pnl_by_day = {}
    day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for e in closed:
        if e.entry_at:
            day = day_names[e.entry_at.weekday()]
            pnl_by_day[day] = round(pnl_by_day.get(day, 0) + (e.pnl_usd or 0), 2)

    # Streaks
    longest_win = longest_loss = cur_win = cur_loss = 0
    for e in closed:
        if e.status == "closed_win":
            cur_win += 1
            cur_loss = 0
            longest_win = max(longest_win, cur_win)
        elif e.status == "closed_loss":
            cur_loss += 1
            cur_win = 0
            longest_loss = max(longest_loss, cur_loss)
        else:
            cur_win = 0
            cur_loss = 0

    return JournalStatsResponse(
        total_trades=len(entries),
        open_trades=len([e for e in entries if e.status == "open"]),
        closed_trades=len(closed),
        wins=len(wins),
        losses=len(losses),
        breakeven=len(be),
        win_rate=round(win_rate, 1),
        total_pnl_usd=round(total_pnl, 2),
        avg_pnl_usd=round(avg_pnl, 2),
        avg_rr=round(avg_rr, 2),
        best_trade_pnl=round(best.pnl_usd, 2) if best and best.pnl_usd else 0,
        worst_trade_pnl=round(worst.pnl_usd, 2) if worst and worst.pnl_usd else 0,
        best_trade_pair=best.pair if best else None,
        worst_trade_pair=worst.pair if worst else None,
        most_traded_pair=most_traded,
        avg_confidence_wins=avg_emotion(wins, "confidence"),
        avg_confidence_losses=avg_emotion(losses, "confidence"),
        avg_fomo_wins=avg_emotion(wins, "fomo_level"),
        avg_fomo_losses=avg_emotion(losses, "fomo_level"),
        most_common_mistake=most_common_mistake,
        most_profitable_strategy=best_strat,
        longest_win_streak=longest_win,
        longest_loss_streak=longest_loss,
        win_rate_by_strategy=wr_by_strat,
        win_rate_by_emotion=wr_by_emotion,
        pnl_by_day=pnl_by_day,
    )


# ════════════════════════════════════════════
# 4. AI INSIGHTS (Gemini 2.5 Flash)
# ════════════════════════════════════════════

@router.get("/ai/insights", response_model=AIInsightResponse)
async def get_ai_insights(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Generate AI coach insights from journal data using Gemini."""
    cache_key = f"lq:journal:insights:{current_user.id}"
    cached = cache_get(cache_key)
    if cached:
        return AIInsightResponse(**cached)

    # Get stats
    stats = await get_journal_stats(current_user, db)

    if stats.total_trades < 3:
        return AIInsightResponse(
            insights=["Need at least 3 journal entries to generate insights. Keep journaling!"],
            generated_at=datetime.now(timezone.utc).isoformat(),
            source="system",
        )

    prompt = f"""You are a professional crypto trading coach analyzing a trader's journal data.
Generate exactly 4 short, actionable insights (max 2 sentences each).
Be specific with numbers. Use $ amounts and percentages.
Focus on: emotional patterns, strategy performance, costly mistakes, and one positive edge.

Trader's data:
- Total trades: {stats.total_trades}, Win rate: {stats.win_rate}%
- Total PnL: ${stats.total_pnl_usd}, Avg PnL: ${stats.avg_pnl_usd}
- Best trade: {stats.best_trade_pair} (${stats.best_trade_pnl}), Worst: {stats.worst_trade_pair} (${stats.worst_trade_pnl})
- Win rate by strategy: {stats.win_rate_by_strategy}
- Win rate by emotion/mood: {stats.win_rate_by_emotion}
- PnL by day of week: {stats.pnl_by_day}
- Most common mistake: {stats.most_common_mistake}
- Avg confidence on wins: {stats.avg_confidence_wins}, on losses: {stats.avg_confidence_losses}
- Avg FOMO on wins: {stats.avg_fomo_wins}, on losses: {stats.avg_fomo_losses}
- Longest win streak: {stats.longest_win_streak}, loss streak: {stats.longest_loss_streak}

Return ONLY a JSON array of 4 strings. No markdown, no explanation.
Example: ["Insight 1 here.", "Insight 2 here.", "Insight 3 here.", "Insight 4 here."]"""

    if not GEMINI_API_KEY:
        return AIInsightResponse(
            insights=["AI insights unavailable — Gemini API key not configured."],
            generated_at=datetime.now(timezone.utc).isoformat(),
            source="system",
        )

    try:
        async with httpx.AsyncClient(timeout=20.0) as client:
            resp = await client.post(
                f"{GEMINI_URL}?key={GEMINI_API_KEY}",
                json={
                    "contents": [{"parts": [{"text": prompt}]}],
                    "generationConfig": {"temperature": 0.7, "maxOutputTokens": 500},
                },
            )

        if resp.status_code != 200:
            return AIInsightResponse(
                insights=[f"AI temporarily unavailable (status {resp.status_code}). Check back later."],
                generated_at=datetime.now(timezone.utc).isoformat(),
                source="error",
            )

        body = resp.json()
        raw_text = body["candidates"][0]["content"]["parts"][0]["text"]

        # Parse JSON array from response
        import json
        clean = raw_text.strip()
        if clean.startswith("```"):
            clean = clean.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
        insights = json.loads(clean)
        if not isinstance(insights, list):
            insights = [str(insights)]

        result = {
            "insights": insights[:5],
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "source": "gemini",
        }
        cache_set(cache_key, result, ttl=900)  # 15 min cache
        return AIInsightResponse(**result)

    except Exception as e:
        return AIInsightResponse(
            insights=[f"Could not generate AI insights: {str(e)[:100]}"],
            generated_at=datetime.now(timezone.utc).isoformat(),
            source="error",
        )


# ════════════════════════════════════════════
# 5. EXCEL EXPORT
# ════════════════════════════════════════════

@router.get("/export/excel")
async def export_journal_excel(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    pair: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """Export journal to styled .xlsx with 4 sheets (trade log, summary, charts, AI insights)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.chart import LineChart, BarChart, PieChart, Reference
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    # Fetch entries
    q = db.query(TradeJournal).filter(TradeJournal.user_id == current_user.id)
    if pair:
        q = q.filter(TradeJournal.pair == pair.upper())
    if status_filter:
        q = q.filter(TradeJournal.status == status_filter)
    if date_from:
        q = q.filter(TradeJournal.entry_at >= date_from)
    if date_to:
        q = q.filter(TradeJournal.entry_at <= date_to)
    entries = q.order_by(TradeJournal.entry_at.asc()).all()

    if not entries:
        raise HTTPException(status_code=404, detail="No journal entries to export")

    wb = Workbook()

    # ── Styles ──
    gold_fill = PatternFill("solid", fgColor="D4A853")
    dark_fill = PatternFill("solid", fgColor="1A0A0C")
    green_fill = PatternFill("solid", fgColor="E8F5E9")
    red_fill = PatternFill("solid", fgColor="FFEBEE")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    data_font = Font(size=10, name="Arial")
    money_fmt = '#,##0.00'
    pct_fmt = '0.00"%"'
    thin_border = Border(
        left=Side(style="thin", color="D4A853"),
        right=Side(style="thin", color="D4A853"),
        top=Side(style="thin", color="D4A853"),
        bottom=Side(style="thin", color="D4A853"),
    )

    # ════════════════════════════════════════
    # SHEET 1: Trade Log
    # ════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Trade Log"
    ws1.sheet_properties.tabColor = "D4A853"

    headers = [
        "Date", "Pair", "Direction", "Status",
        "Planned Entry", "Actual Entry", "Actual Exit",
        "Leverage", "Size (USD)", "Fees (USD)",
        "PnL (USD)", "PnL (%)", "R:R",
        "Strategy", "Confluences", "Mood", "Confidence",
        "FOMO", "Mistakes", "Signal Linked", "Notes",
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = gold_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, e in enumerate(entries, 2):
        row_data = [
            e.entry_at.strftime("%Y-%m-%d %H:%M") if e.entry_at else "",
            e.pair,
            e.direction.upper(),
            e.status.replace("_", " ").title(),
            e.planned_entry,
            e.actual_entry,
            e.actual_exit,
            e.leverage,
            e.position_size_usd,
            e.fees_usd,
            e.pnl_usd,
            e.pnl_pct,
            e.rr_ratio,
            ", ".join(e.strategy_tags or []),
            ", ".join(e.confluence_tags or []),
            (e.emotions or {}).get("mood", ""),
            (e.emotions or {}).get("confidence", ""),
            (e.emotions or {}).get("fomo_level", ""),
            ", ".join(e.mistakes or []),
            "Yes" if e.signal_id else "No",
            (e.notes or "")[:100],
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in (5, 6, 7, 9, 10, 11):
                cell.number_format = money_fmt
            if col_idx == 12:
                cell.number_format = '0.00'

        # Conditional fill
        pnl = e.pnl_usd
        if pnl is not None and pnl > 0:
            for c in range(1, len(headers) + 1):
                ws1.cell(row=row_idx, column=c).fill = green_fill
        elif pnl is not None and pnl < 0:
            for c in range(1, len(headers) + 1):
                ws1.cell(row=row_idx, column=c).fill = red_fill

    # Auto width
    for col_idx in range(1, len(headers) + 1):
        max_len = max(len(str(ws1.cell(row=r, column=col_idx).value or "")) for r in range(1, len(entries) + 2))
        ws1.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    ws1.auto_filter.ref = ws1.dimensions
    ws1.freeze_panes = "A2"

    # ════════════════════════════════════════
    # SHEET 2: Summary Stats
    # ════════════════════════════════════════
    ws2 = wb.create_sheet("Summary Stats")
    ws2.sheet_properties.tabColor = "534AB7"

    stats = await get_journal_stats(current_user, db, date_from, date_to)

    summary_data = [
        ("LuxQuant Trade Journal Summary", ""),
        ("", ""),
        ("Total Trades", stats.total_trades),
        ("Wins", stats.wins),
        ("Losses", stats.losses),
        ("Win Rate", f"{stats.win_rate}%"),
        ("", ""),
        ("Total PnL (USD)", stats.total_pnl_usd),
        ("Avg PnL per Trade", stats.avg_pnl_usd),
        ("Avg Risk:Reward", stats.avg_rr),
        ("", ""),
        ("Best Trade", f"{stats.best_trade_pair} (${stats.best_trade_pnl})"),
        ("Worst Trade", f"{stats.worst_trade_pair} (${stats.worst_trade_pnl})"),
        ("Most Traded Pair", stats.most_traded_pair or "N/A"),
        ("", ""),
        ("Avg Confidence (Wins)", stats.avg_confidence_wins or "N/A"),
        ("Avg Confidence (Losses)", stats.avg_confidence_losses or "N/A"),
        ("Avg FOMO (Wins)", stats.avg_fomo_wins or "N/A"),
        ("Avg FOMO (Losses)", stats.avg_fomo_losses or "N/A"),
        ("", ""),
        ("Most Common Mistake", stats.most_common_mistake or "None"),
        ("Most Profitable Strategy", stats.most_profitable_strategy or "N/A"),
        ("Longest Win Streak", stats.longest_win_streak),
        ("Longest Loss Streak", stats.longest_loss_streak),
    ]

    for row_idx, (label, val) in enumerate(summary_data, 1):
        lc = ws2.cell(row=row_idx, column=1, value=label)
        vc = ws2.cell(row=row_idx, column=2, value=val)
        lc.font = Font(bold=True, size=11, name="Arial") if row_idx == 1 else Font(size=10, name="Arial", color="666666")
        vc.font = Font(bold=True, size=11, name="Arial") if row_idx == 1 else Font(size=10, name="Arial")
        if row_idx == 1:
            lc.font = Font(bold=True, size=14, color="D4A853", name="Arial")

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 25

    # ════════════════════════════════════════
    # SHEET 3: Charts
    # ════════════════════════════════════════
    ws3 = wb.create_sheet("Charts")
    ws3.sheet_properties.tabColor = "BA7517"

    # Chart data: Equity curve
    ws3.cell(row=1, column=1, value="Trade #").font = Font(bold=True, size=10, name="Arial")
    ws3.cell(row=1, column=2, value="Cumulative PnL").font = Font(bold=True, size=10, name="Arial")
    closed_entries = [e for e in entries if e.status != "open"]
    running = 0
    for i, e in enumerate(closed_entries, 2):
        running += (e.pnl_usd or 0)
        ws3.cell(row=i, column=1, value=i - 1)
        ws3.cell(row=i, column=2, value=round(running, 2))

    if len(closed_entries) >= 2:
        chart1 = LineChart()
        chart1.title = "Equity Curve"
        chart1.x_axis.title = "Trade #"
        chart1.y_axis.title = "Cumulative PnL ($)"
        chart1.style = 10
        chart1.width = 20
        chart1.height = 12
        data_ref = Reference(ws3, min_col=2, min_row=1, max_row=len(closed_entries) + 1)
        cats = Reference(ws3, min_col=1, min_row=2, max_row=len(closed_entries) + 1)
        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.series[0].graphicalProperties.line.solidFill = "4ade80"
        ws3.add_chart(chart1, "D1")

    # Chart data: Win rate by strategy (start at col 5)
    strat_data = stats.win_rate_by_strategy
    if strat_data:
        start_row_s = 1
        ws3.cell(row=start_row_s, column=5, value="Strategy").font = Font(bold=True, size=10, name="Arial")
        ws3.cell(row=start_row_s, column=6, value="Win Rate %").font = Font(bold=True, size=10, name="Arial")
        for i, (strat, info) in enumerate(strat_data.items(), start_row_s + 1):
            ws3.cell(row=i, column=5, value=strat)
            ws3.cell(row=i, column=6, value=info["win_rate"])

        chart2 = BarChart()
        chart2.type = "bar"
        chart2.title = "Win Rate by Strategy"
        chart2.x_axis.title = "Strategy"
        chart2.y_axis.title = "Win Rate %"
        chart2.style = 10
        chart2.width = 18
        chart2.height = 10
        data_ref2 = Reference(ws3, min_col=6, min_row=start_row_s, max_row=start_row_s + len(strat_data))
        cats2 = Reference(ws3, min_col=5, min_row=start_row_s + 1, max_row=start_row_s + len(strat_data))
        chart2.add_data(data_ref2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.series[0].graphicalProperties.solidFill = "D4A853"
        ws3.add_chart(chart2, "D16")

    # Chart data: PnL by day
    pnl_day = stats.pnl_by_day
    if pnl_day:
        day_order = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        start_col = 8
        ws3.cell(row=1, column=start_col, value="Day").font = Font(bold=True, size=10, name="Arial")
        ws3.cell(row=1, column=start_col + 1, value="PnL ($)").font = Font(bold=True, size=10, name="Arial")
        row_i = 2
        for day in day_order:
            if day in pnl_day:
                ws3.cell(row=row_i, column=start_col, value=day)
                ws3.cell(row=row_i, column=start_col + 1, value=pnl_day[day])
                row_i += 1

        if row_i > 2:
            chart3 = BarChart()
            chart3.title = "PnL by Day of Week"
            chart3.style = 10
            chart3.width = 18
            chart3.height = 10
            data_ref3 = Reference(ws3, min_col=start_col + 1, min_row=1, max_row=row_i - 1)
            cats3 = Reference(ws3, min_col=start_col, min_row=2, max_row=row_i - 1)
            chart3.add_data(data_ref3, titles_from_data=True)
            chart3.set_categories(cats3)
            chart3.series[0].graphicalProperties.solidFill = "D4A853"
            ws3.add_chart(chart3, "D31")

    # Win/Loss pie
    ws3.cell(row=1, column=11, value="Result").font = Font(bold=True, size=10, name="Arial")
    ws3.cell(row=1, column=12, value="Count").font = Font(bold=True, size=10, name="Arial")
    ws3.cell(row=2, column=11, value="Wins")
    ws3.cell(row=2, column=12, value=stats.wins)
    ws3.cell(row=3, column=11, value="Losses")
    ws3.cell(row=3, column=12, value=stats.losses)

    if stats.wins + stats.losses > 0:
        pie = PieChart()
        pie.title = "Win vs Loss"
        pie.style = 10
        pie.width = 14
        pie.height = 10
        pie_data = Reference(ws3, min_col=12, min_row=1, max_row=3)
        pie_cats = Reference(ws3, min_col=11, min_row=2, max_row=3)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(pie_cats)
        ws3.add_chart(pie, "N1")

    # ════════════════════════════════════════
    # SHEET 4: AI Insights
    # ════════════════════════════════════════
    ws4 = wb.create_sheet("AI Insights")
    ws4.sheet_properties.tabColor = "378ADD"

    ws4.cell(row=1, column=1, value="LuxQuant AI Coach Insights").font = Font(bold=True, size=14, color="D4A853", name="Arial")
    ws4.cell(row=2, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").font = Font(size=9, color="999999", name="Arial")
    ws4.cell(row=3, column=1, value="AI-generated analysis based on your journal data.").font = Font(size=9, color="999999", name="Arial")

    try:
        ai_resp = await get_ai_insights(current_user, db)
        for i, insight in enumerate(ai_resp.insights, 5):
            ws4.cell(row=i, column=1, value=f"• {insight}").font = Font(size=11, name="Arial")
            ws4.row_dimensions[i].height = 30
    except Exception:
        ws4.cell(row=5, column=1, value="AI insights not available — view in LuxQuant dashboard.").font = Font(size=11, name="Arial")

    ws4.column_dimensions["A"].width = 90

    # ── Save to buffer ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"LuxQuant_Journal_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )